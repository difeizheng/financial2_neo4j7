"""Excel parser — reads workbook into CellNode list."""

from __future__ import annotations

from openpyxl import load_workbook

from models.cell_node import CellNode, col_to_index


def col_letter(col_idx: int) -> str:
    """Convert 1-based column index to letter."""
    chars = []
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        chars.append(chr(65 + rem))
    return "".join(reversed(chars))


def parse_workbook(path: str) -> list[CellNode]:
    """Extract all non-empty cells from an Excel file as CellNode list."""
    wb = load_workbook(path, data_only=False)  # data_only=False to get formulas
    cells: list[CellNode] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect merged cell ranges
        merged_ranges = set()
        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_ranges.add((row, col))

        # Collect cells with values or formulas
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                col = col_letter(cell.column)
                cell_id = f"{sheet_name}_{cell.row}_{col}"

                is_formula = cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("=")
                is_merged = (cell.row, cell.column) in merged_ranges

                # Determine data type
                if is_formula:
                    data_type = "formula"
                elif isinstance(cell.value, bool):
                    data_type = "boolean"
                elif isinstance(cell.value, (int, float)):
                    data_type = "number"
                elif isinstance(cell.value, str):
                    data_type = "text"
                else:
                    data_type = "text"

                # Merge range for merged cells
                merge_range = None
                if is_merged:
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row == cell.row and mr.min_col == cell.column:
                            merge_range = str(mr)
                            break

                node = CellNode(
                    id=cell_id,
                    sheet=sheet_name,
                    row=cell.row,
                    col=col,
                    value=cell.value if not is_formula else None,
                    formula_raw=cell.value if is_formula else None,
                    data_type=data_type,
                    is_merged=is_merged,
                    merge_range=merge_range,
                    format_code=cell.number_format if hasattr(cell, "number_format") else None,
                )
                cells.append(node)

    return cells


def parse_workbook_with_values(path: str) -> list[CellNode]:
    """Also get the computed values (data_only=True) alongside formulas."""
    cells = parse_workbook(path)

    # Get values
    wb = load_workbook(path, data_only=True)
    value_map = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col = col_letter(cell.column)
                    cell_id = f"{sheet_name}_{cell.row}_{col}"
                    value_map[cell_id] = cell.value

    # Update cells with values
    for cell in cells:
        if cell.formula_raw and cell.id in value_map:
            cell.value = value_map[cell.id]

    return cells
